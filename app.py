from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime, date
import os
import sqlite3
import csv
import io

from flask import Response

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


app = Flask(__name__)
app.secret_key = 'secretkey'

# ======================
# FIXED DATABASE PATH
# ======================
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
INSTANCE_DIR = os.path.join(BASE_DIR, 'instance')

os.makedirs(INSTANCE_DIR, exist_ok=True)

DB_PATH = os.path.join(INSTANCE_DIR, 'ccs.db')

app.config['UPLOAD_FOLDER'] = os.path.join(BASE_DIR, 'static', 'uploads')
app.config['SQLALCHEMY_DATABASE_URI'] = f"sqlite:///{DB_PATH}"
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

TIME_SLOTS = [
    '07:00 AM – 09:00 AM',
    '09:00 AM – 11:00 AM',
    '11:00 AM – 01:00 PM',
    '01:00 PM – 03:00 PM',
    '03:00 PM – 05:00 PM',
    '05:00 PM – 07:00 PM',
]

LABS = ['Lab 524', 'Lab 526', 'Lab 528', 'Lab 530', 'Lab 542', 'Lab 544']
PCS_PER_LAB = 40


# ======================
# MODELS
# ======================

class Student(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    id_number = db.Column(db.String(50), unique=True, nullable=False)
    last_name = db.Column(db.String(100))
    first_name = db.Column(db.String(100))
    middle_name = db.Column(db.String(100))
    course = db.Column(db.String(50))
    course_level = db.Column(db.String(10))
    email = db.Column(db.String(120), unique=True)
    password = db.Column(db.String(200))
    address = db.Column(db.String(200))
    profile_pic = db.Column(db.String(200), default='default.png')
    remaining_sessions = db.Column(db.Integer, default=30)

    leaderboard_points = db.Column(db.Integer, default=0)
    raw_points = db.Column(db.Float, default=0)
    cleanliness_points = db.Column(db.Integer, default=0)
    hours_points = db.Column(db.Integer, default=0)
    task_points = db.Column(db.Integer, default=0)
    total_hours = db.Column(db.Float, default=0)


class Admin(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True)
    password = db.Column(db.String(200))


class Announcement(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    author = db.Column(db.String(100), default='CCS Admin')
    content = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.String(50))


class SitInRecord(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_id_number = db.Column(db.String(50))
    student_name = db.Column(db.String(150))
    purpose = db.Column(db.String(100))
    sit_lab = db.Column(db.String(50))
    session = db.Column(db.String(50))
    status = db.Column(db.String(20), default='Active')
    login_time = db.Column(
        db.String(50),
        default=lambda: datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    )
    logout_time = db.Column(db.String(50), nullable=True)


class Feedback(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('student.id'), nullable=False)
    sitin_record_id = db.Column(db.Integer, db.ForeignKey('sit_in_record.id'), nullable=False)
    rating = db.Column(db.Integer, nullable=False)
    feedback_text = db.Column(db.Text, nullable=False)
    created_at = db.Column(
        db.String(50),
        default=lambda: datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    )

    student = db.relationship('Student', backref='feedbacks')
    sitin_record = db.relationship('SitInRecord', backref='feedback_entries')


class Reservation(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('student.id'), nullable=False)
    student_name = db.Column(db.String(150), nullable=False)
    student_id_number = db.Column(db.String(50), nullable=False)

    lab = db.Column(db.String(50), nullable=False)
    date = db.Column(db.String(20), nullable=False)
    time_slot = db.Column(db.String(100), nullable=False)
    pc_number = db.Column(db.Integer, nullable=False)

    status = db.Column(db.String(20), default='Pending')
    created_at = db.Column(
        db.String(50),
        default=lambda: datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    )

    student = db.relationship('Student', backref='reservations')


class PCStatus(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    lab = db.Column(db.String(50), nullable=False)
    pc_number = db.Column(db.Integer, nullable=False)
    status = db.Column(db.String(20), default='Available')

    __table_args__ = (
        db.UniqueConstraint('lab', 'pc_number', name='unique_lab_pc'),
    )


# ======================
# DATABASE HELPERS
# ======================

def _db_path():
    return DB_PATH


def ensure_student_remaining_sessions_column():
    db_path = _db_path()
    if not os.path.exists(db_path):
        return

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    cursor.execute("PRAGMA table_info(student)")
    columns = [col[1] for col in cursor.fetchall()]

    if 'remaining_sessions' not in columns:
        cursor.execute("ALTER TABLE student ADD COLUMN remaining_sessions INTEGER DEFAULT 30")
        conn.commit()

    conn.close()


def ensure_sitin_logout_columns():
    db_path = _db_path()
    if not os.path.exists(db_path):
        return

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    cursor.execute("PRAGMA table_info(sit_in_record)")
    columns = [col[1] for col in cursor.fetchall()]

    if 'login_time' not in columns:
        cursor.execute("ALTER TABLE sit_in_record ADD COLUMN login_time TEXT")
    if 'logout_time' not in columns:
        cursor.execute("ALTER TABLE sit_in_record ADD COLUMN logout_time TEXT")

    conn.commit()
    conn.close()


def ensure_feedback_columns():
    db_path = _db_path()
    if not os.path.exists(db_path):
        return

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='feedback'")
    table_exists = cursor.fetchone()

    if not table_exists:
        conn.close()
        return

    cursor.execute("PRAGMA table_info(feedback)")
    columns = [col[1] for col in cursor.fetchall()]

    if 'student_id' not in columns:
        cursor.execute("ALTER TABLE feedback ADD COLUMN student_id INTEGER")
    if 'sitin_record_id' not in columns:
        cursor.execute("ALTER TABLE feedback ADD COLUMN sitin_record_id INTEGER")
    if 'rating' not in columns:
        cursor.execute("ALTER TABLE feedback ADD COLUMN rating INTEGER DEFAULT 5")
    if 'feedback_text' not in columns:
        cursor.execute("ALTER TABLE feedback ADD COLUMN feedback_text TEXT")
    if 'created_at' not in columns:
        cursor.execute("ALTER TABLE feedback ADD COLUMN created_at TEXT")

    conn.commit()
    conn.close()

def ensure_leaderboard_columns():
    db_path = _db_path()

    if not os.path.exists(db_path):
        return

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    cursor.execute("PRAGMA table_info(student)")
    columns = [col[1] for col in cursor.fetchall()]

    leaderboard_columns = {
    'leaderboard_points': 'INTEGER DEFAULT 0',
    'raw_points': 'REAL DEFAULT 0',
    'cleanliness_points': 'INTEGER DEFAULT 0',
    'hours_points': 'INTEGER DEFAULT 0',
    'task_points': 'INTEGER DEFAULT 0',
    'total_hours': 'REAL DEFAULT 0'
}

    for column, column_type in leaderboard_columns.items():
        if column not in columns:
            cursor.execute(
                f"ALTER TABLE student ADD COLUMN {column} {column_type}"
            )

    conn.commit()
    conn.close()


def ensure_pc_status_seed():
    for lab in LABS:
        for pc_number in range(1, PCS_PER_LAB + 1):
            existing = PCStatus.query.filter_by(lab=lab, pc_number=pc_number).first()
            if not existing:
                db.session.add(PCStatus(lab=lab, pc_number=pc_number, status='Available'))
    db.session.commit()


def get_full_name(student):
    return f"{student.first_name} {student.middle_name or ''} {student.last_name}".replace("  ", " ").strip()


def is_valid_date_string(date_str):
    try:
        datetime.strptime(date_str, '%Y-%m-%d')
        return True
    except ValueError:
        return False


def get_reserved_pc_numbers(lab, reservation_date, time_slot):
    reservations = Reservation.query.filter(
        Reservation.lab == lab,
        Reservation.date == reservation_date,
        Reservation.time_slot == time_slot,
        Reservation.status.in_(['Pending', 'Approved'])
    ).all()
    return {r.pc_number for r in reservations}


def get_maintenance_pc_numbers(lab):
    pcs = PCStatus.query.filter_by(lab=lab, status='Maintenance').all()
    return {pc.pc_number for pc in pcs}


def get_available_pc_numbers(lab, reservation_date, time_slot):
    reserved = get_reserved_pc_numbers(lab, reservation_date, time_slot)
    maintenance = get_maintenance_pc_numbers(lab)

    available = []
    for pc_number in range(1, PCS_PER_LAB + 1):
        if pc_number not in reserved and pc_number not in maintenance:
            available.append(pc_number)
    return available


def init_database():
    with app.app_context():
        db.create_all()
        ensure_student_remaining_sessions_column()
        ensure_sitin_logout_columns()
        ensure_feedback_columns()
        ensure_leaderboard_columns()
        ensure_pc_status_seed()

        students = Student.query.all()
        for student in students:
            if student.remaining_sessions is None:
                student.remaining_sessions = 30

        if not Admin.query.filter_by(username='admin').first():
            admin = Admin(
                username='admin',
                password=generate_password_hash('admin123')
            )
            db.session.add(admin)

        db.session.commit()


# ======================
# HOME
# ======================

@app.route('/')
def home():
    return redirect(url_for('login'))


# ======================
# LOGIN
# ======================

@app.route('/login', methods=['GET', 'POST'])
def login():
    top_students = Student.query.order_by(
        Student.leaderboard_points.desc()
    ).limit(3).all()

    if request.method == 'POST':
        id_number = request.form.get('id_number', '').strip()
        password = request.form.get('password', '')

        admin = Admin.query.filter_by(username=id_number).first()

        if admin and check_password_hash(admin.password, password):
            session.clear()
            session['admin_id'] = admin.id
            flash('Login successful!', 'success')
            return redirect(url_for('admin_dashboard'))

        student = Student.query.filter_by(id_number=id_number).first()

        if not student:
            flash('Account not found.', 'error')
        elif not check_password_hash(student.password, password):
            flash('Incorrect password.', 'error')
        else:
            session.clear()
            session['student_id'] = student.id
            flash('Login successful!', 'success')
            return redirect(url_for('dashboard'))

    return render_template(
        'login.html',
        top_students=top_students
    )


# ======================
# REGISTER
# ======================

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        id_number = request.form.get('id_number', '').strip()
        last_name = request.form.get('last_name', '').strip()
        first_name = request.form.get('first_name', '').strip()
        middle_name = request.form.get('middle_name', '').strip()
        course = request.form.get('course', '').strip()
        course_level = request.form.get('course_level', '').strip()
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')
        repeat_password = request.form.get('repeat_password', '')
        address = request.form.get('address', '').strip()

        form_data = {
            'id_number': id_number,
            'last_name': last_name,
            'first_name': first_name,
            'middle_name': middle_name,
            'course': course,
            'course_level': course_level,
            'email': email,
            'address': address
        }

        if not all([id_number, last_name, first_name, course, course_level, email, password, repeat_password]):
            flash('Please fill in all required fields.', 'error')
            return render_template('register.html', form_data=form_data)

        if password != repeat_password:
            flash('Passwords do not match.', 'error')
            return render_template('register.html', form_data=form_data)

        if Student.query.filter_by(id_number=id_number).first():
            flash('ID already exists.', 'error')
            return render_template('register.html', form_data=form_data)

        if Student.query.filter_by(email=email).first():
            flash('Email already exists.', 'error')
            return render_template('register.html', form_data=form_data)

        hashed_pw = generate_password_hash(password)

        new_student = Student(
            id_number=id_number,
            last_name=last_name,
            first_name=first_name,
            middle_name=middle_name,
            course=course,
            course_level=course_level,
            email=email,
            password=hashed_pw,
            address=address,
            remaining_sessions=30
        )

        db.session.add(new_student)
        db.session.commit()

        flash('Registered successfully!', 'success')
        return redirect(url_for('login'))

    return render_template('register.html', form_data={})


# ======================
# ADMIN DASHBOARD
# ======================

@app.route('/admin_dashboard')
def admin_dashboard():
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    students = Student.query.all()
    total_students = Student.query.count()
    announcements = Announcement.query.order_by(Announcement.id.desc()).all()
    current_sitin_count = SitInRecord.query.filter_by(status='Active').count()
    total_sitin_count = SitInRecord.query.count()

    purpose_rows = db.session.query(
        SitInRecord.purpose,
        db.func.count(SitInRecord.id)
    ).group_by(SitInRecord.purpose).all()

    purpose_labels = [row[0] for row in purpose_rows]
    purpose_counts = [row[1] for row in purpose_rows]

    return render_template(
        'admin_dashboard.html',
        students=students,
        total_students=total_students,
        announcements=announcements,
        current_sitin_count=current_sitin_count,
        total_sitin_count=total_sitin_count,
        purpose_labels=purpose_labels,
        purpose_counts=purpose_counts
    )


# ======================
# STUDENT DASHBOARD
# ======================

@app.route('/dashboard')
def dashboard():

    if 'student_id' not in session:
        return redirect(url_for('login'))

    student = Student.query.get_or_404(
        session['student_id']
    )

    announcements, notif_reservations, feedback_notifications = \
        get_student_notification_data(student)

    return render_template(
        'dashboard.html',
        student=student,
        announcements=announcements,
        notif_reservations=notif_reservations,
        feedback_notifications=feedback_notifications
    )

# ======================
# HISTORY
# ======================

@app.route('/history')
def history():

    if 'student_id' not in session:
        return redirect(url_for('login'))

    student = Student.query.get_or_404(
        session['student_id']
    )

    records = SitInRecord.query.filter_by(
        student_id_number=student.id_number
    ).order_by(
        SitInRecord.id.desc()
    ).all()

    feedbacks = Feedback.query.filter_by(
        student_id=student.id
    ).all()

    submitted_feedback_ids = {
        feedback.sitin_record_id
        for feedback in feedbacks
    }

    total_history = len(records)

    logged_out_count = sum(
        1 for record in records
        if record.status == 'Logged Out'
    )

    pending_feedback_count = sum(
        1 for record in records
        if record.status == 'Logged Out'
        and record.id not in submitted_feedback_ids
    )

    submitted_feedback_count = len(feedbacks)

    announcements, notif_reservations, feedback_notifications = \
        get_student_notification_data(student)

    return render_template(
        'history.html',
        student=student,
        records=records,
        submitted_feedback_ids=submitted_feedback_ids,
        total_history=total_history,
        logged_out_count=logged_out_count,
        pending_feedback_count=pending_feedback_count,
        submitted_feedback_count=submitted_feedback_count,
        announcements=announcements,
        notif_reservations=notif_reservations,
        feedback_notifications=feedback_notifications
    )

# ======================
# ANNOUNCEMENT SUBMIT
# ======================

@app.route('/submit_announcement', methods=['POST'])
def submit_announcement():
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    content = request.form.get('content', '').strip()

    if content:
        new_announcement = Announcement(
            content=content,
            created_at=datetime.now().strftime('%Y-%b-%d')
        )
        db.session.add(new_announcement)
        db.session.commit()
        flash('Announcement posted successfully!', 'success')
    else:
        flash('Announcement cannot be empty.', 'error')

    return redirect(url_for('admin_dashboard'))


# ======================
# STUDENTS PAGE
# ======================

@app.route('/admin_students')
def admin_students():
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    students = Student.query.order_by(Student.id.desc()).all()
    return render_template('admin_students.html', students=students)


@app.route('/add_student', methods=['POST'])
def add_student():
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    id_number = request.form.get('id_number', '').strip()
    first_name = request.form.get('first_name', '').strip()
    middle_name = request.form.get('middle_name', '').strip()
    last_name = request.form.get('last_name', '').strip()
    course = request.form.get('course', '').strip()
    course_level = request.form.get('course_level', '').strip()
    email = request.form.get('email', '').strip()
    address = request.form.get('address', '').strip()

    if not all([id_number, first_name, last_name, course, course_level, email]):
        flash('Please fill all required fields.', 'error')
        return redirect(url_for('admin_students'))

    if Student.query.filter_by(id_number=id_number).first():
        flash('Student ID already exists.', 'error')
        return redirect(url_for('admin_students'))

    if Student.query.filter_by(email=email).first():
        flash('Email already exists.', 'error')
        return redirect(url_for('admin_students'))

    default_password = generate_password_hash('123456')

    new_student = Student(
        id_number=id_number,
        first_name=first_name,
        middle_name=middle_name,
        last_name=last_name,
        course=course,
        course_level=course_level,
        email=email,
        address=address,
        password=default_password,
        remaining_sessions=30
    )

    db.session.add(new_student)
    db.session.commit()

    flash('Student added successfully.', 'success')
    return redirect(url_for('admin_students'))


@app.route('/edit_student/<int:student_id>', methods=['POST'])
def edit_student(student_id):
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    student = Student.query.get_or_404(student_id)

    id_number = request.form.get('id_number', '').strip()
    first_name = request.form.get('first_name', '').strip()
    middle_name = request.form.get('middle_name', '').strip()
    last_name = request.form.get('last_name', '').strip()
    course = request.form.get('course', '').strip()
    course_level = request.form.get('course_level', '').strip()
    email = request.form.get('email', '').strip()
    address = request.form.get('address', '').strip()
    remaining_sessions = request.form.get('remaining_sessions', '').strip()

    if not all([id_number, first_name, last_name, course, course_level, email]):
        flash('Please fill all required fields.', 'error')
        return redirect(url_for('admin_students'))

    existing_id = Student.query.filter(
        Student.id_number == id_number,
        Student.id != student.id
    ).first()
    if existing_id:
        flash('Another student already uses that ID number.', 'error')
        return redirect(url_for('admin_students'))

    existing_email = Student.query.filter(
        Student.email == email,
        Student.id != student.id
    ).first()
    if existing_email:
        flash('Another student already uses that email.', 'error')
        return redirect(url_for('admin_students'))

    try:
        remaining_sessions = int(remaining_sessions)
    except ValueError:
        flash('Remaining sessions must be a valid number.', 'error')
        return redirect(url_for('admin_students'))

    student.id_number = id_number
    student.first_name = first_name
    student.middle_name = middle_name
    student.last_name = last_name
    student.course = course
    student.course_level = course_level
    student.email = email
    student.address = address
    student.remaining_sessions = remaining_sessions

    db.session.commit()
    flash('Student updated successfully.', 'success')
    return redirect(url_for('admin_students'))


@app.route('/delete_student/<int:student_id>', methods=['POST'])
def delete_student(student_id):
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    student = Student.query.get_or_404(student_id)

    active_record = SitInRecord.query.filter_by(
        student_id_number=student.id_number,
        status='Active'
    ).first()

    if active_record:
        flash('Cannot delete a student with an active sit-in.', 'error')
        return redirect(url_for('admin_students'))

    Feedback.query.filter_by(student_id=student.id).delete()
    SitInRecord.query.filter_by(student_id_number=student.id_number).delete()
    Reservation.query.filter_by(student_id=student.id).delete()

    db.session.delete(student)
    db.session.commit()

    flash('Student deleted successfully.', 'success')
    return redirect(url_for('admin_students'))


@app.route('/reset_all_sessions', methods=['POST'])
def reset_all_sessions():
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    students = Student.query.all()
    for student in students:
        student.remaining_sessions = 30

    db.session.commit()
    flash('All student sessions have been reset to 30.', 'success')
    return redirect(url_for('admin_students'))


# ======================
# CURRENT SIT-IN
# ======================

@app.route('/current_sitin')
def current_sitin():
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    records = SitInRecord.query.filter_by(
        status='Active'
    ).order_by(SitInRecord.id.desc()).all()

    students = Student.query.all()

    student_pic_map = {
        student.id_number: student.profile_pic
        for student in students
    }

    return render_template(
        'current_sitin.html',
        sitin_records=records,
        student_pic_map=student_pic_map
    )


@app.route('/view_sitin_records')
def view_sitin_records():
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    records = SitInRecord.query.order_by(SitInRecord.id.desc()).all()
    return render_template('view_sitin_records.html', sitin_records=records)


@app.route('/submit_sitin', methods=['POST'])
def submit_sitin():
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    student_id_number = request.form.get('student_id_number')
    purpose = request.form.get('purpose')
    sit_lab = request.form.get('sit_lab')

    student = Student.query.filter_by(id_number=student_id_number).first()

    if not student:
        flash('Student not found.', 'error')
        return redirect(url_for('admin_dashboard'))

    if student.remaining_sessions <= 0:
        flash('This student has no remaining sessions.', 'error')
        return redirect(url_for('admin_dashboard'))

    active_record = SitInRecord.query.filter_by(
        student_id_number=student.id_number,
        status='Active'
    ).first()

    if active_record:
        flash('This student already has an active sit-in.', 'error')
        return redirect(url_for('current_sitin'))

    record = SitInRecord(
        student_id_number=student.id_number,
        student_name=get_full_name(student),
        purpose=purpose,
        sit_lab=sit_lab,
        session=str(student.remaining_sessions),
        status='Active',
        login_time=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    )

    db.session.add(record)
    db.session.commit()

    flash('Student successfully sat in!', 'success')
    return redirect(url_for('current_sitin'))

@app.route('/logout_sitin/<int:record_id>', methods=['POST'])
def logout_sitin(record_id):

    if 'admin_id' not in session:
        return redirect(url_for('login'))

    record = SitInRecord.query.get_or_404(record_id)

    if record.status == 'Logged Out':
        flash('This sit-in record is already logged out.', 'error')
        return redirect(url_for('view_sitin_records'))

    student = Student.query.filter_by(
        id_number=record.student_id_number
    ).first()

    if not student:
        flash('Student not found.', 'error')
        return redirect(url_for('current_sitin'))

    clean_workstation = request.form.get('clean_workstation')
    total_hours_checked = request.form.get('total_hours')
    task_completed = request.form.get('task_completed')

    logout_datetime = datetime.now()

    try:
        login_datetime = datetime.strptime(
            record.login_time,
            '%Y-%m-%d %H:%M:%S'
        )

        duration_seconds = (
            logout_datetime - login_datetime
        ).total_seconds()

        total_minutes = int(duration_seconds // 60)

        if total_minutes < 0:
            total_minutes = 0

        hours = total_minutes // 60
        minutes = total_minutes % 60
        duration_hours = round(total_minutes / 60, 2)

    except Exception:
        total_minutes = 0
        hours = 0
        minutes = 0
        duration_hours = 0

    student.total_hours = round(
        (student.total_hours or 0) + duration_hours,
        2
    )

    earned_score = 0
    earned_raw_point = 0
    checked_count = 0

    if clean_workstation:
        student.cleanliness_points += 5
        earned_score += 5
        checked_count += 1

    if total_hours_checked:
        student.hours_points += 3
        earned_score += 3
        checked_count += 1

    if task_completed:
        student.task_points += 2
        earned_score += 2
        checked_count += 1

    student.leaderboard_points += earned_score

    if checked_count == 3:
        earned_raw_point = 1
    elif checked_count == 2:
        earned_raw_point = 0.50
    elif checked_count == 1:
        earned_raw_point = 0.25

    previous_raw_points = student.raw_points or 0
    student.raw_points = previous_raw_points + earned_raw_point

    previous_rewards = int(previous_raw_points // 3)
    current_rewards = int(student.raw_points // 3)
    new_rewards = current_rewards - previous_rewards

    if new_rewards > 0:
        student.remaining_sessions += new_rewards
        flash(
            f'Student earned +{new_rewards} extra session(s)!',
            'success'
        )

    if student.remaining_sessions > 0:
        student.remaining_sessions -= 1

    record.session = str(student.remaining_sessions)
    record.status = 'Logged Out'
    record.logout_time = logout_datetime.strftime('%Y-%m-%d %H:%M:%S')

    db.session.commit()

    flash(
        f'Student logged out successfully! '
        f'Session time: {hours}h {minutes}m, '
        f'+{earned_score} leaderboard score, '
        f'+{earned_raw_point} raw point.',
        'success'
    )

    return redirect(url_for('view_sitin_records'))

# ======================
# RESERVATION - STUDENT
# ======================

@app.route('/reservation')
@app.route('/student/reservation')
def reservation():
    if 'student_id' not in session:
        return redirect(url_for('login'))

    student = Student.query.get_or_404(session['student_id'])

    reservations = Reservation.query.filter_by(
        student_id=student.id
    ).order_by(Reservation.id.desc()).all()

    announcements, notif_reservations, pending_feedback_records = \
        get_student_notification_data(student)

    return render_template(
        'reservation.html',
        reservations=reservations,
        student=student,
        announcements=announcements,
        notif_reservations=notif_reservations,
        pending_feedback_records=pending_feedback_records
    )


@app.route('/submit_reservation', methods=['POST'])
def submit_reservation():
    if 'student_id' not in session:
        return redirect(url_for('login'))

    student = Student.query.get_or_404(session['student_id'])

    reservation_date = request.form.get('date', '').strip()
    lab = request.form.get('lab', '').strip()
    time_slot = request.form.get('time_slot', '').strip()
    pc_number = request.form.get('pc_number', '').strip()

    if not all([reservation_date, lab, time_slot, pc_number]):
        flash('Please complete the reservation details.', 'error')
        return redirect(url_for('reservation'))

    if not is_valid_date_string(reservation_date):
        flash('Invalid reservation date.', 'error')
        return redirect(url_for('reservation'))

    if reservation_date < date.today().strftime('%Y-%m-%d'):
        flash('You cannot reserve a past date.', 'error')
        return redirect(url_for('reservation'))

    if lab not in LABS:
        flash('Invalid laboratory selected.', 'error')
        return redirect(url_for('reservation'))

    if time_slot not in TIME_SLOTS:
        flash('Invalid time slot selected.', 'error')
        return redirect(url_for('reservation'))

    # BLOCK TIME SLOTS THAT ALREADY STARTED/PASSED TODAY
    if reservation_date == date.today().strftime('%Y-%m-%d'):
        slot_start_part = time_slot.split('–')[0].strip()
        slot_start_time = datetime.strptime(slot_start_part, '%I:%M %p').time()

        if slot_start_time <= datetime.now().time():
            flash('You cannot reserve a time slot that already started or passed.', 'error')
            return redirect(url_for('reservation'))

    try:
        pc_number = int(pc_number)
    except ValueError:
        flash('Invalid PC number.', 'error')
        return redirect(url_for('reservation'))

    if pc_number < 1 or pc_number > PCS_PER_LAB:
        flash('Invalid PC number.', 'error')
        return redirect(url_for('reservation'))

    maintenance_pcs = get_maintenance_pc_numbers(lab)
    if pc_number in maintenance_pcs:
        flash('That PC is under maintenance.', 'error')
        return redirect(url_for('reservation'))

    existing_pc_reservation = Reservation.query.filter(
        Reservation.lab == lab,
        Reservation.date == reservation_date,
        Reservation.time_slot == time_slot,
        Reservation.pc_number == pc_number,
        Reservation.status.in_(['Pending', 'Approved'])
    ).first()

    if existing_pc_reservation:
        flash('That PC is already reserved for the selected schedule.', 'error')
        return redirect(url_for('reservation'))

    existing_student_slot = Reservation.query.filter(
        Reservation.student_id == student.id,
        Reservation.date == reservation_date,
        Reservation.time_slot == time_slot,
        Reservation.status.in_(['Pending', 'Approved'])
    ).first()

    if existing_student_slot:
        flash('You already have a reservation for that date and time slot.', 'error')
        return redirect(url_for('reservation'))

    new_reservation = Reservation(
        student_id=student.id,
        student_name=get_full_name(student),
        student_id_number=student.id_number,
        lab=lab,
        date=reservation_date,
        time_slot=time_slot,
        pc_number=pc_number,
        status='Pending'
    )

    db.session.add(new_reservation)
    db.session.commit()

    flash('Reservation submitted successfully!', 'success')
    return redirect(url_for('reservation'))

@app.route('/cancel_reservation/<int:reservation_id>', methods=['POST'])
def cancel_reservation(reservation_id):
    if 'student_id' not in session:
        return redirect(url_for('login'))

    reservation = Reservation.query.get_or_404(reservation_id)

    if reservation.student_id != session['student_id']:
        flash('You are not allowed to cancel this reservation.', 'error')
        return redirect(url_for('reservation'))

    if reservation.status == 'Cancelled':
        flash('Reservation already cancelled.', 'error')
        return redirect(url_for('reservation'))

    reservation.status = 'Cancelled'
    db.session.commit()

    flash('Reservation cancelled successfully.', 'success')
    return redirect(url_for('reservation'))


# ======================
# RESERVATION - ADMIN
# ======================

@app.route('/admin/reservations')
@app.route('/admin_reservation')
def admin_reservation():
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    reservations = Reservation.query.order_by(Reservation.id.desc()).all()
    return render_template('admin_reservation.html', reservations=reservations)


@app.route('/admin/approve_reservation/<int:reservation_id>', methods=['POST'])
def approve_reservation(reservation_id):
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    reservation = Reservation.query.get_or_404(reservation_id)

    if reservation.status == 'Cancelled':
        flash('Cancelled reservation cannot be approved.', 'error')
        return redirect(url_for('admin_reservation'))

    existing_conflict = Reservation.query.filter(
        Reservation.id != reservation.id,
        Reservation.lab == reservation.lab,
        Reservation.date == reservation.date,
        Reservation.time_slot == reservation.time_slot,
        Reservation.pc_number == reservation.pc_number,
        Reservation.status == 'Approved'
    ).first()

    if existing_conflict:
        flash('Cannot approve. PC is already approved for that schedule.', 'error')
        return redirect(url_for('admin_reservation'))

    if reservation.pc_number in get_maintenance_pc_numbers(reservation.lab):
        flash('Cannot approve. Selected PC is under maintenance.', 'error')
        return redirect(url_for('admin_reservation'))

    reservation.status = 'Approved'
    db.session.commit()

    flash('Reservation approved.', 'success')
    return redirect(url_for('admin_reservation'))

@app.route('/leaderboard')
def leaderboard():
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    students = Student.query.order_by(
        Student.leaderboard_points.desc()
    ).limit(10).all()

    return render_template(
        'leaderboard.html',
        students=students
    )


@app.route('/admin/decline_reservation/<int:reservation_id>', methods=['POST'])
def decline_reservation(reservation_id):
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    reservation = Reservation.query.get_or_404(reservation_id)
    reservation.status = 'Cancelled'
    db.session.commit()

    flash('Reservation declined.', 'success')
    return redirect(url_for('admin_reservation'))


@app.route('/admin/cancel_reservation/<int:reservation_id>', methods=['POST'])
def cancel_reservation_admin(reservation_id):
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    reservation = Reservation.query.get_or_404(reservation_id)
    reservation.status = 'Cancelled'
    db.session.commit()

    flash('Reservation cancelled by admin.', 'success')
    return redirect(url_for('admin_reservation'))

# ======================
# REAL-TIME RESERVATION API
# ======================

@app.route('/api/reservation/labs')
def api_reservation_labs():
    if 'student_id' not in session and 'admin_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    reservation_date = request.args.get('date', '').strip()

    if not reservation_date or not is_valid_date_string(reservation_date):
        return jsonify({'error': 'Valid date is required'}), 400

    labs_data = []
    for lab in LABS:
        maintenance_count = len(get_maintenance_pc_numbers(lab))
        usable_pcs = PCS_PER_LAB - maintenance_count

        labs_data.append({
            'id': lab,
            'name': lab,
            'pcs': PCS_PER_LAB,
            'available_count': usable_pcs,
            'status': 'Available' if usable_pcs > 0 else 'Full'
        })

    return jsonify({'labs': labs_data})


@app.route('/api/reservation/time-slots')
def api_reservation_time_slots():

    if 'student_id' not in session and 'admin_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    reservation_date = request.args.get('date', '').strip()
    lab = request.args.get('lab', '').strip()

    if not reservation_date or not is_valid_date_string(reservation_date):
        return jsonify({'error': 'Valid date is required'}), 400

    if lab not in LABS:
        return jsonify({'error': 'Valid lab is required'}), 400

    slots = []

    for slot in TIME_SLOTS:

        available_pcs = get_available_pc_numbers(
            lab,
            reservation_date,
            slot
        )

        # CHECK IF SLOT IS ALREADY PAST
        is_past_slot = False

        if reservation_date == date.today().strftime('%Y-%m-%d'):

            slot_start_part = slot.split('–')[0].strip()
            slot_start_time = datetime.strptime(
                slot_start_part,
                '%I:%M %p'
            ).time()

            if slot_start_time <= datetime.now().time():
                is_past_slot = True

        slots.append({
            'time_slot': slot,
            'available_count': len(available_pcs),
            'is_full': len(available_pcs) == 0,
            'is_past': is_past_slot
        })

    return jsonify({'time_slots': slots})


@app.route('/api/reservation/pcs')
def api_reservation_pcs():
    if 'student_id' not in session and 'admin_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    reservation_date = request.args.get('date', '').strip()
    lab = request.args.get('lab', '').strip()
    time_slot = request.args.get('time_slot', '').strip()

    if not reservation_date or not is_valid_date_string(reservation_date):
        return jsonify({'error': 'Valid date is required'}), 400

    if lab not in LABS:
        return jsonify({'error': 'Valid lab is required'}), 400

    if time_slot not in TIME_SLOTS:
        return jsonify({'error': 'Valid time slot is required'}), 400

    reserved = get_reserved_pc_numbers(lab, reservation_date, time_slot)
    maintenance = get_maintenance_pc_numbers(lab)

    pcs = []
    for pc_number in range(1, PCS_PER_LAB + 1):
        if pc_number in maintenance:
            status = 'Maintenance'
        elif pc_number in reserved:
            status = 'Reserved'
        else:
            status = 'Available'

        pcs.append({
            'pc_number': pc_number,
            'status': status
        })

    return jsonify({'pcs': pcs})


# ======================
# ADMIN PC STATUS API
# ======================

@app.route('/api/admin/pc-status')
def api_admin_pc_status():
    if 'admin_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    lab = request.args.get('lab', '').strip()
    if lab not in LABS:
        return jsonify({'error': 'Valid lab is required'}), 400

    pcs = PCStatus.query.filter_by(lab=lab).order_by(PCStatus.pc_number.asc()).all()

    return jsonify({
        'pcs': [
            {
                'pc_number': pc.pc_number,
                'status': pc.status
            }
            for pc in pcs
        ]
    })


@app.route('/api/admin/toggle-pc-status', methods=['POST'])
def api_admin_toggle_pc_status():
    if 'admin_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    lab = request.form.get('lab', '').strip()
    pc_number = request.form.get('pc_number', '').strip()

    if lab not in LABS:
        return jsonify({'error': 'Valid lab is required'}), 400

    try:
        pc_number = int(pc_number)
    except ValueError:
        return jsonify({'error': 'Invalid PC number'}), 400

    pc = PCStatus.query.filter_by(lab=lab, pc_number=pc_number).first()
    if not pc:
        return jsonify({'error': 'PC not found'}), 404

    pc.status = 'Available' if pc.status == 'Maintenance' else 'Maintenance'
    db.session.commit()

    return jsonify({
        'success': True,
        'pc_number': pc.pc_number,
        'status': pc.status
    })


# ======================
# ADMIN NAVBAR ALIASES
# ======================

@app.route('/admin_sitin')
def admin_sitin():
    return redirect(url_for('current_sitin'))


@app.route('/admin_records')
def admin_records():
    return redirect(url_for('view_sitin_records'))


@app.route('/admin_feedback')
def admin_feedback():
    return redirect(url_for('feedback_reports'))


# ======================
# FEEDBACK
# ======================

@app.route('/submit_feedback/<int:record_id>', methods=['POST'])
def submit_feedback(record_id):
    if 'student_id' not in session:
        return redirect(url_for('login'))

    student = Student.query.get_or_404(session['student_id'])
    record = SitInRecord.query.get_or_404(record_id)

    if record.student_id_number != student.id_number:
        flash('You are not allowed to submit feedback for this record.', 'error')
        return redirect(url_for('history'))

    if record.status != 'Logged Out':
        flash('You can only submit feedback after logout.', 'error')
        return redirect(url_for('history'))

    existing_feedback = Feedback.query.filter_by(
        student_id=student.id,
        sitin_record_id=record.id
    ).first()

    if existing_feedback:
        flash('Feedback already submitted for this sit-in record.', 'error')
        return redirect(url_for('history'))

    rating = request.form.get('rating', '').strip()
    feedback_text = request.form.get('feedback_text', '').strip()

    if not rating:
        flash('Please select a rating.', 'error')
        return redirect(url_for('history'))

    try:
        rating = int(rating)
    except ValueError:
        flash('Invalid rating value.', 'error')
        return redirect(url_for('history'))

    if rating < 1 or rating > 5:
        flash('Rating must be between 1 and 5.', 'error')
        return redirect(url_for('history'))

    if not feedback_text:
        flash('Feedback cannot be empty.', 'error')
        return redirect(url_for('history'))

    new_feedback = Feedback(
        student_id=student.id,
        sitin_record_id=record.id,
        rating=rating,
        feedback_text=feedback_text
    )

    db.session.add(new_feedback)
    db.session.commit()

    flash('Feedback submitted successfully.', 'success')
    return redirect(url_for('history'))


@app.route('/feedback_reports')
def feedback_reports():
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    feedbacks = Feedback.query.order_by(Feedback.id.desc()).all()
    total_feedbacks = len(feedbacks)

    ratings = [item.rating for item in feedbacks]
    overall_rating = round(sum(ratings) / total_feedbacks, 1) if total_feedbacks else 0
    avg_rating = overall_rating
    highest_rating = max(ratings) if ratings else 0
    lowest_rating = min(ratings) if ratings else 0

    rating_counts = {5: 0, 4: 0, 3: 0, 2: 0, 1: 0}
    for item in feedbacks:
        if item.rating in rating_counts:
            rating_counts[item.rating] += 1

    return render_template(
        'feedback_reports.html',
        feedbacks=feedbacks,
        total_feedbacks=total_feedbacks,
        overall_rating=overall_rating,
        rating_counts=rating_counts,
        avg_rating=avg_rating,
        highest_rating=highest_rating,
        lowest_rating=lowest_rating
    )


# ======================
# LOGOUT
# ======================

@app.route('/logout')
def logout():
    session.clear()
    flash('Logged out successfully.', 'success')
    return redirect(url_for('login'))


@app.route('/admin_logout')
def admin_logout():
    session.clear()
    flash('Admin logged out.', 'success')
    return redirect(url_for('login'))


# ======================
# EDIT PROFILE
# ======================

@app.route('/edit_profile', methods=['GET', 'POST'])
def edit_profile():
    if 'student_id' not in session:
        return redirect(url_for('login'))

    student = Student.query.get_or_404(session['student_id'])

    announcements, notif_reservations, pending_feedback_records = \
        get_student_notification_data(student)

    if request.method == 'POST':
        student.first_name = request.form.get('first_name')
        student.last_name = request.form.get('last_name')
        student.middle_name = request.form.get('middle_name')
        student.email = request.form.get('email')
        student.address = request.form.get('address')

        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')

        if new_password or confirm_password:
            if new_password != confirm_password:
                flash('Passwords do not match.', 'error')
                return render_template(
                    'edit_profile.html',
                    student=student,
                    announcements=announcements,
                    notif_reservations=notif_reservations,
                    pending_feedback_records=pending_feedback_records
                )

            if len(new_password) < 6:
                flash('Password must be at least 6 characters.', 'error')
                return render_template(
                    'edit_profile.html',
                    student=student,
                    announcements=announcements,
                    notif_reservations=notif_reservations,
                    pending_feedback_records=pending_feedback_records
                )

            student.password = generate_password_hash(new_password)

        file = request.files.get('profile_pic')

        if file and file.filename:
            filename = secure_filename(file.filename)
            os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)
            student.profile_pic = filename

        db.session.commit()
        flash('Profile updated successfully!', 'success')
        return redirect(url_for('dashboard'))

    return render_template(
        'edit_profile.html',
        student=student,
        announcements=announcements,
        notif_reservations=notif_reservations,
        pending_feedback_records=pending_feedback_records
    )


# ======================
# FORGOT PASSWORD
# ======================

@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        id_number = request.form.get('id_number', '').strip()
        email = request.form.get('email', '').strip()

        student = Student.query.filter_by(id_number=id_number, email=email).first()

        if not student:
            flash('ID number and email do not match any account.', 'error')
            return render_template('forgot_password.html')

        return redirect(url_for('reset_password', student_id=student.id))

    return render_template('forgot_password.html')


@app.route('/reset_password/<int:student_id>', methods=['GET', 'POST'])
def reset_password(student_id):
    student = Student.query.get_or_404(student_id)

    if request.method == 'POST':
        new_password = request.form.get('new_password', '')
        confirm_password = request.form.get('confirm_password', '')

        if not new_password or not confirm_password:
            flash('Please fill in both password fields.', 'error')
            return render_template('reset_password.html', student=student)

        if new_password != confirm_password:
            flash('Passwords do not match.', 'error')
            return render_template('reset_password.html', student=student)

        if len(new_password) < 6:
            flash('Password must be at least 6 characters.', 'error')
            return render_template('reset_password.html', student=student)

        student.password = generate_password_hash(new_password)
        db.session.commit()

        flash('Password reset successfully. You can now log in.', 'success')
        return redirect(url_for('login'))

    return render_template('reset_password.html', student=student)


# ======================
# DEBUG ROUTE
# ======================

@app.route('/debug_reservations')
def debug_reservations():
    reservations = Reservation.query.order_by(Reservation.id.desc()).all()
    return {
        'db_path': DB_PATH,
        'count': len(reservations),
        'items': [
            {
                'id': r.id,
                'student_name': r.student_name,
                'student_id_number': r.student_id_number,
                'lab': r.lab,
                'date': r.date,
                'time_slot': r.time_slot,
                'pc_number': r.pc_number,
                'status': r.status
            }
            for r in reservations
        ]
    }

def get_filtered_sitin_records():
    query = SitInRecord.query

    search = request.args.get('search', '').strip()
    purpose = request.args.get('purpose', '').strip()
    lab = request.args.get('lab', '').strip()
    date_filter = request.args.get('date', '').strip()

    if search:
        search_like = f'%{search}%'
        query = query.filter(
            db.or_(
                SitInRecord.student_name.ilike(search_like),
                SitInRecord.student_id_number.ilike(search_like),
                SitInRecord.purpose.ilike(search_like),
                SitInRecord.sit_lab.ilike(search_like),
                SitInRecord.status.ilike(search_like)
            )
        )

    if purpose:
        query = query.filter(SitInRecord.purpose == purpose)

    if lab:
        query = query.filter(SitInRecord.sit_lab == lab)

    if date_filter:
        query = query.filter(SitInRecord.login_time.like(f'{date_filter}%'))

    return query.order_by(SitInRecord.id.desc()).all()

@app.route('/sitin_reports')
def sitin_reports():
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    sit_in_records = SitInRecord.query.order_by(
        SitInRecord.id.desc()
    ).all()

    return render_template(
        'sitin_reports.html',
        sit_in_records=sit_in_records
    )

@app.route('/export_csv')
def export_csv():

    if 'admin_id' not in session:
        return redirect(url_for('login'))

    records = get_filtered_sitin_records()

    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow([
        'ID Number',
        'Student Name',
        'Purpose',
        'Laboratory',
        'Login Time',
        'Logout Time',
        'Status'
    ])

    for r in records:
        writer.writerow([
            r.student_id_number,
            r.student_name,
            r.purpose,
            r.sit_lab,
            r.login_time,
            r.logout_time if r.logout_time else '',
            r.status
        ])

    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={
            'Content-Disposition':
            'attachment; filename=sitin_reports.csv'
        }
    )

@app.route('/export_excel')
def export_excel():

    if 'admin_id' not in session:
        return redirect(url_for('login'))

    records = get_filtered_sitin_records()

    workbook = Workbook()
    sheet = workbook.active

    sheet.title = "Sit-in Reports"

    headers = [
        'ID Number',
        'Student Name',
        'Purpose',
        'Laboratory',
        'Login Time',
        'Logout Time',
        'Status'
    ]

    sheet.append(headers)

    header_fill = PatternFill(
        start_color="4B2D8F",
        end_color="4B2D8F",
        fill_type="solid"
    )

    for cell in sheet[1]:
        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = header_fill

        cell.alignment = Alignment(
            horizontal="center"
        )

    for r in records:
        sheet.append([
            r.student_id_number,
            r.student_name,
            r.purpose,
            r.sit_lab,
            r.login_time,
            r.logout_time if r.logout_time else '',
            r.status
        ])

    for column_cells in sheet.columns:

        max_length = 0

        column_letter = column_cells[0].column_letter

        for cell in column_cells:

            try:
                value = str(cell.value)

                if len(value) > max_length:
                    max_length = len(value)

            except:
                pass

        adjusted_width = max_length + 4

        sheet.column_dimensions[
            column_letter
        ].width = adjusted_width

    output = io.BytesIO()

    workbook.save(output)

    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition':
            'attachment; filename=sitin_reports.xlsx'
        }
    )


   # =========================================
# STUDENT NOTIFICATION HELPER
# =========================================

def get_student_notification_data(student):

    announcements = Announcement.query.order_by(
        Announcement.id.desc()
    ).limit(5).all()

    notif_reservations = Reservation.query.filter_by(
        student_id=student.id
    ).order_by(
        Reservation.id.desc()
    ).limit(5).all()

    records = SitInRecord.query.filter_by(
        student_id_number=student.id_number
    ).order_by(
        SitInRecord.id.desc()
    ).limit(5).all()

    feedbacks = Feedback.query.filter_by(
        student_id=student.id
    ).all()

    submitted_feedback_ids = {
        feedback.sitin_record_id
        for feedback in feedbacks
    }

    feedback_notifications = []

    for record in records:
        if record.status == 'Logged Out':
            feedback_notifications.append({
                'record': record,
                'submitted': record.id in submitted_feedback_ids
            })

    return (
        announcements,
        notif_reservations,
        feedback_notifications
    )

init_database()

if __name__ == '__main__':
    app.run(debug=True)